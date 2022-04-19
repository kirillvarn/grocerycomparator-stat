from openpyxl import Workbook
from openpyxl.styles import Font


class Excel:
    def __init__(self, filename="data"):
        self.filename = filename
        self.workbook = Workbook()

    def append_header(self, header: list, sheet_name: str = "Sheet"):
        ws = self.workbook[sheet_name]
        ws.append(header)

    def create_sheets(self, sheet_names, **optional):
        del self.workbook["Sheet"]
        for name in sheet_names:
            self.workbook.create_sheet(name)

            if 'addition_sheet_name' in optional:
                self.workbook.create_sheet(
                    f"{name} {optional['addition_sheet_name']}")

    def put_data(self, data: list, sheet_name: str = "Sheet"):
        ws = self.workbook[sheet_name]
        colors = {
            "up": Font(size=11, bold=True, color='002ef06b'),
            "down": Font(size=11, bold=True, color='00e34f32')
        }

        for row in range(0, len(data)):
            for col in range(0, len(data[row])):
                trend = "still"
                if isinstance(data[row][col], dict):
                    value = data[row][col]["value"]
                    trend = data[row][col]["trend"]
                else:
                    value = data[row][col]

                ws.cell(column=col+1, row=row+2, value=value)
                if trend != "still":
                    ws.cell(row+2, col+1).font = colors[trend]

        # for values in data:
        #     ws.append(values)
    def put_regression_data(data):
        pass

    def save(self):
        self.workbook.save(f"{self.filename}.xlsx")
